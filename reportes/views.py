from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import FileResponse, HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from .models import Reporte
from .forms import ReporteForm
import json
import os
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame
from reportlab.lib.units import inch, cm
from balanza.models import Balanza
from liquidaciones.models import Liquidacion, LiquidacionDetalle
from decimal import Decimal
from django.conf import settings
from reportlab.lib.utils import ImageReader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.core.files.base import ContentFile
from io import BytesIO
from lotes.models import Campana, CampanaLote, Lote
from django.views.decorators.http import require_POST

# Vista que muestra la lista de reportes generados por el usuario actual
@login_required
def reporte_list(request):
    # Ahora que los reportes se descargan directamente, solo mostramos el formulario de creación
    if request.method == 'POST':
        form = ReporteForm(request.POST)
        if form.is_valid():
            reporte = form.save(commit=False)
            reporte.usuario_creador = request.user
            
            # Generar y descargar el reporte según el tipo
            if reporte.tipo == 'lotes':
                return generar_reporte_lotes_excel_directo(reporte)
            elif reporte.tipo == 'leyes':
                return generar_reporte_leyes_directo(reporte)
            elif reporte.tipo == 'costos':
                return generar_reporte_costos_directo(reporte)
            elif reporte.tipo == 'valorizaciones':
                return generar_reporte_valorizaciones_directo(reporte)
            elif reporte.tipo == 'liquidaciones':
                return generar_reporte_liquidaciones_excel_directo(reporte)
            elif reporte.tipo == 'balanza':
                return generar_reporte_balanza_excel_directo(reporte)
            elif reporte.tipo == 'campanas':
                return generar_reporte_campanas_excel_directo(reporte)
    else:
        form = ReporteForm()
    
    return render(request, 'reportes/reporte_list.html', {'form': form})

# Vista para crear un nuevo reporte Excel según el tipo seleccionado
@login_required
def reporte_create(request):
    if request.method == 'POST':
        form = ReporteForm(request.POST)
        if form.is_valid():
            reporte = form.save(commit=False)
            reporte.usuario_creador = request.user
            
            # Generar y descargar el reporte según el tipo
            if reporte.tipo == 'lotes':
                return generar_reporte_lotes_excel_directo(reporte)
            elif reporte.tipo == 'leyes':
                return generar_reporte_leyes_directo(reporte)
            elif reporte.tipo == 'costos':
                return generar_reporte_costos_directo(reporte)
            elif reporte.tipo == 'valorizaciones':
                return generar_reporte_valorizaciones_directo(reporte)
            elif reporte.tipo == 'liquidaciones':
                return generar_reporte_liquidaciones_excel_directo(reporte)
            elif reporte.tipo == 'balanza':
                return generar_reporte_balanza_excel_directo(reporte)
            elif reporte.tipo == 'campanas':
                return generar_reporte_campanas_excel_directo(reporte)
    else:
        form = ReporteForm()
    
    return render(request, 'reportes/reporte_form.html', {'form': form})

# Vista para descargar el archivo Excel de un reporte generado
@login_required
def reporte_download(request, pk):
    reporte = get_object_or_404(Reporte, pk=pk, usuario_creador=request.user)
    if reporte.archivo:
        return FileResponse(reporte.archivo, as_attachment=True)
    messages.error(request, 'El archivo del reporte no existe.')
    return redirect('reportes:reporte_list')

# Vista para eliminar un reporte generado por el usuario
@require_POST
@login_required
def reporte_delete(request, pk):
    reporte = get_object_or_404(Reporte, pk=pk, usuario_creador=request.user)
    reporte.delete()
    messages.success(request, 'Reporte eliminado correctamente.')
    return redirect('reportes:reporte_list')

# Función para generar y descargar directamente el reporte de lotes en Excel
def generar_reporte_lotes_excel_directo(reporte):
    from lotes.models import Lote, CampanaLote
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lotes'

    # Encabezados
    headers = [
        'Código Lote', 'Fecha Ingreso', 'TMH', 'TMS Real', 'Tipo Producto', 'Facturador', 'Transportista', 'Estado', 'Campaña Asociada'
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Filtrar lotes por fechas
    lotes = Lote.objects.filter(
        fecha_ingreso__gte=reporte.fecha_inicio,
        fecha_ingreso__lte=reporte.fecha_fin
    ).select_related('costo', 'tipo_producto', 'facturador', 'transportista').order_by('fecha_ingreso')

    for lote in lotes:
        # Buscar campaña asociada (puede ser None)
        campana_lote = CampanaLote.objects.filter(lote=lote).select_related('campana').first()
        campana_nombre = campana_lote.campana.nombre if campana_lote else ''
        costo = getattr(lote, 'costo', None)
        ws.append([
            lote.codigo_lote,
            lote.fecha_ingreso.strftime('%Y-%m-%d'),
            lote.tmh,
            getattr(costo, 'tms_real', '-') if costo else '-',
            getattr(lote.tipo_producto, 'nombre', ''),
            getattr(lote.facturador, 'razon_social', ''),
            getattr(lote.transportista, 'razon_social', ''),
            getattr(lote, 'estado', ''),
            campana_nombre
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Generar respuesta HTTP directa
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_lotes_{reporte.fecha_inicio}_{reporte.fecha_fin}.xlsx"'
    
    # Guardar en memoria y enviar
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.read())
    
    return response

# Función para generar y descargar directamente el reporte de liquidaciones en Excel
def generar_reporte_liquidaciones_excel_directo(reporte):
    from liquidaciones.models import Liquidacion, LiquidacionDetalle
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Liquidaciones'

    # Encabezados
    headers = [
        'ID', 'Fecha', 'Proveedor', 'RUC', 'N° Lotes', 'Total Valorización', 'IGV', 'Anticipo', 'Detracción', 'Descuento Flete', 'Monto Pagado'
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Filtrar liquidaciones por fechas
    liquidaciones = Liquidacion.objects.filter(
        fecha_creacion__gte=reporte.fecha_inicio,
        fecha_creacion__lte=reporte.fecha_fin
    ).select_related('proveedor').order_by('fecha_creacion')

    for liq in liquidaciones:
        detalles = LiquidacionDetalle.objects.filter(liquidacion=liq)
        subtotal = sum(det.lote.valorizacion.valorizacion_uds_tms or 0 for det in detalles)
        igv = subtotal * 0.18
        total_valorizacion = subtotal + igv
        anticipo = sum(det.lote.valorizacion.anticipo or 0 for det in detalles)
        total_valorizacion_neto = total_valorizacion - anticipo
        detraccion = total_valorizacion_neto * 0.10
        descuento_flete = sum(
            (det.lote.tmh or Decimal('0')) * (getattr(det.lote.valorizacion, 'pu_tmh_flete', Decimal('0')) or Decimal('0'))
            for det in detalles
        )
        monto_pagado = total_valorizacion_neto - detraccion - descuento_flete
        ws.append([
            liq.id,
            liq.fecha_creacion.strftime('%Y-%m-%d'),
            getattr(liq.proveedor, 'razon_social', ''),
            getattr(liq.proveedor, 'ruc', ''),
            detalles.count(),
            round(total_valorizacion, 2),
            round(igv, 2),
            round(anticipo, 2),
            round(detraccion, 2),
            round(descuento_flete, 2),
            round(monto_pagado, 2)
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Generar respuesta HTTP directa
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_liquidaciones_{reporte.fecha_inicio}_{reporte.fecha_fin}.xlsx"'
    
    # Guardar en memoria y enviar
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.read())
    
    return response

# Función para generar y descargar directamente el reporte de balanza en Excel
def generar_reporte_balanza_excel_directo(reporte):
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Balanza'

    # Encabezados
    headers = [
        'Fecha', 'Guía Ticket', 'Facturador', 'Vehículo', 'Conductor',
        'Producto', 'Peso Ingreso (kg)', 'Peso Salida (kg)', 'Peso Neto (kg)',
        'Guía Remisión', 'Guía Transporte', 'Observaciones'
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Filtrar por fechas
    balanzas = Balanza.objects.filter(
        fecha_ingreso__gte=reporte.fecha_inicio,
        fecha_ingreso__lte=reporte.fecha_fin
    ).order_by('fecha_ingreso')

    for b in balanzas:
        ws.append([
            b.fecha_ingreso.strftime('%Y-%m-%d'),
            b.numero_guia_ticket,
            getattr(b.facturador, 'razon_social', ''),
            getattr(b.vehiculo, 'placa', ''),
            getattr(b.conductor, 'nombres', ''),
            getattr(b.tipo_producto, 'nombre', ''),
            b.peso_ingreso_kg,
            b.peso_salida_kg,
            (b.peso_ingreso_kg or 0) - (b.peso_salida_kg or 0),
            b.guia_remision,
            b.guia_transporte,
            b.observaciones
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Generar respuesta HTTP directa
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_balanza_{reporte.fecha_inicio}_{reporte.fecha_fin}.xlsx"'
    
    # Guardar en memoria y enviar
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.read())
    
    return response

# Función para generar y descargar directamente el reporte de campañas en Excel
def generar_reporte_campanas_excel_directo(reporte):
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Campañas'

    # Encabezados
    headers = [
        'Nombre', 'Descripción', 'Fecha Inicio', 'Fecha Fin', 'Estado', 'Lotes Asociados', 'Total TMH', 'Total TMS Real'
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Filtrar campañas por fechas
    campanas = Campana.objects.filter(
        fecha_inicio__gte=reporte.fecha_inicio,
        fecha_fin__lte=reporte.fecha_fin
    ).order_by('fecha_inicio')

    for campana in campanas:
        lotes = Lote.objects.filter(campanalote__campana=campana).select_related('costo')
        lotes_str = ', '.join([lote.codigo_lote for lote in lotes])
        total_tmh = sum(lote.tmh or 0 for lote in lotes)
        total_tms = sum(getattr(lote.costo, 'tms_real', 0) or 0 for lote in lotes)
        ws.append([
            campana.nombre,
            campana.descripcion or '',
            campana.fecha_inicio.strftime('%Y-%m-%d'),
            campana.fecha_fin.strftime('%Y-%m-%d'),
            campana.get_estado_display(),
            lotes_str,
            total_tmh,
            total_tms
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Generar respuesta HTTP directa
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_campanas_{reporte.fecha_inicio}_{reporte.fecha_fin}.xlsx"'
    
    # Guardar en memoria y enviar
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.read())
    
    return response

# Funciones placeholder para los otros tipos de reportes
def generar_reporte_leyes_directo(reporte):
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="reporte_leyes_{reporte.fecha_inicio}_{reporte.fecha_fin}.txt"'
    response.write("Reporte de leyes - Funcionalidad en desarrollo")
    return response

def generar_reporte_costos_directo(reporte):
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="reporte_costos_{reporte.fecha_inicio}_{reporte.fecha_fin}.txt"'
    response.write("Reporte de costos - Funcionalidad en desarrollo")
    return response

def generar_reporte_valorizaciones_directo(reporte):
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="reporte_valorizaciones_{reporte.fecha_inicio}_{reporte.fecha_fin}.txt"'
    response.write("Reporte de valorizaciones - Funcionalidad en desarrollo")
    return response

# Función para generar el reporte de lotes en Excel (mantener para compatibilidad)
def generar_reporte_lotes_excel(reporte):
    from lotes.models import Lote, CampanaLote
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lotes'

    # Encabezados
    headers = [
        'Código Lote', 'Fecha Ingreso', 'TMH', 'TMS Real', 'Tipo Producto', 'Facturador', 'Transportista', 'Estado', 'Campaña Asociada'
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Filtrar lotes por fechas
    lotes = Lote.objects.filter(
        fecha_ingreso__gte=reporte.fecha_inicio,
        fecha_ingreso__lte=reporte.fecha_fin
    ).select_related('costo', 'tipo_producto', 'facturador', 'transportista').order_by('fecha_ingreso')

    for lote in lotes:
        # Buscar campaña asociada (puede ser None)
        campana_lote = CampanaLote.objects.filter(lote=lote).select_related('campana').first()
        campana_nombre = campana_lote.campana.nombre if campana_lote else ''
        costo = getattr(lote, 'costo', None)
        ws.append([
            lote.codigo_lote,
            lote.fecha_ingreso.strftime('%Y-%m-%d'),
            lote.tmh,
            getattr(costo, 'tms_real', '-') if costo else '-',
            getattr(lote.tipo_producto, 'nombre', ''),
            getattr(lote.facturador, 'razon_social', ''),
            getattr(lote.transportista, 'razon_social', ''),
            getattr(lote, 'estado', ''),
            campana_nombre
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Guardar en memoria y asociar al reporte
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    reporte.archivo.save(f'reporte_lotes_{reporte.fecha_inicio}_{reporte.fecha_fin}.xlsx', ContentFile(output.read()), save=False)
    from lotes.models import Lote, CampanaLote
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lotes'

    # Encabezados
    headers = [
        'Código Lote', 'Fecha Ingreso', 'TMH', 'TMS Real', 'Tipo Producto', 'Facturador', 'Transportista', 'Estado', 'Campaña Asociada'
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Filtrar lotes por fechas
    lotes = Lote.objects.filter(
        fecha_ingreso__gte=reporte.fecha_inicio,
        fecha_ingreso__lte=reporte.fecha_fin
    ).select_related('costo', 'tipo_producto', 'facturador', 'transportista').order_by('fecha_ingreso')

    for lote in lotes:
        # Buscar campaña asociada (puede ser None)
        campana_lote = CampanaLote.objects.filter(lote=lote).select_related('campana').first()
        campana_nombre = campana_lote.campana.nombre if campana_lote else ''
        costo = getattr(lote, 'costo', None)
        ws.append([
            lote.codigo_lote,
            lote.fecha_ingreso.strftime('%Y-%m-%d'),
            lote.tmh,
            getattr(costo, 'tms_real', '-') if costo else '-',
            getattr(lote.tipo_producto, 'nombre', ''),
            getattr(lote.facturador, 'razon_social', ''),
            getattr(lote.transportista, 'razon_social', ''),
            getattr(lote, 'estado', ''),
            campana_nombre
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Guardar en memoria y asociar al reporte
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    reporte.archivo.save(f'reporte_lotes_{reporte.fecha_inicio}_{reporte.fecha_fin}.xlsx', ContentFile(output.read()), save=False)

def generar_reporte_leyes(reporte):
    # Implementar la generación del reporte de leyes
    pass

def generar_reporte_costos(reporte):
    # Implementar la generación del reporte de costos
    pass

def generar_reporte_valorizaciones(reporte):
    # Implementar la generación del reporte de valorizaciones
    pass

# Función para generar el reporte de liquidaciones en Excel
def generar_reporte_liquidaciones_excel(reporte):
    from liquidaciones.models import Liquidacion, LiquidacionDetalle
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Liquidaciones'

    # Encabezados
    headers = [
        'ID', 'Fecha', 'Proveedor', 'RUC', 'N° Lotes', 'Total Valorización', 'IGV', 'Anticipo', 'Detracción', 'Descuento Flete', 'Monto Pagado'
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Filtrar liquidaciones por fechas
    liquidaciones = Liquidacion.objects.filter(
        fecha_creacion__gte=reporte.fecha_inicio,
        fecha_creacion__lte=reporte.fecha_fin
    ).select_related('proveedor').order_by('fecha_creacion')

    for liq in liquidaciones:
        detalles = LiquidacionDetalle.objects.filter(liquidacion=liq)
        subtotal = sum(det.lote.valorizacion.valorizacion_uds_tms or Decimal('0') for det in detalles)
        igv = subtotal * Decimal('0.18')
        total_valorizacion = subtotal + igv
        anticipo = sum(det.lote.valorizacion.anticipo or Decimal('0') for det in detalles)
        total_valorizacion_neto = total_valorizacion - anticipo
        detraccion = total_valorizacion_neto * Decimal('0.10')
        descuento_flete = sum(
            (det.lote.tmh or Decimal('0')) * (getattr(det.lote.valorizacion, 'pu_tmh_flete', Decimal('0')) or Decimal('0'))
            for det in detalles
        )
        monto_pagado = total_valorizacion_neto - detraccion - descuento_flete
        ws.append([
            liq.id,
            liq.fecha_creacion.strftime('%Y-%m-%d'),
            getattr(liq.proveedor, 'razon_social', ''),
            getattr(liq.proveedor, 'ruc', ''),
            detalles.count(),
            round(total_valorizacion, 2),
            round(igv, 2),
            round(anticipo, 2),
            round(detraccion, 2),
            round(descuento_flete, 2),
            round(monto_pagado, 2)
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Guardar en memoria y asociar al reporte
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    reporte.archivo.save(f'reporte_liquidaciones_{reporte.fecha_inicio}_{reporte.fecha_fin}.xlsx', ContentFile(output.read()), save=False)

# Función para generar el reporte de balanza en Excel
def generar_reporte_balanza_excel(reporte):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Balanza'

    # Encabezados
    headers = [
        'Fecha', 'Guía Ticket', 'Facturador', 'Vehículo', 'Conductor',
        'Producto', 'Peso Ingreso (kg)', 'Peso Salida (kg)', 'Peso Neto (kg)',
        'Guía Remisión', 'Guía Transporte', 'Observaciones'
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Filtrar por fechas
    balanzas = Balanza.objects.filter(
        fecha_ingreso__gte=reporte.fecha_inicio,
        fecha_ingreso__lte=reporte.fecha_fin
    ).order_by('fecha_ingreso')

    for b in balanzas:
        ws.append([
            b.fecha_ingreso.strftime('%Y-%m-%d'),
            b.numero_guia_ticket,
            getattr(b.facturador, 'razon_social', ''),
            getattr(b.vehiculo, 'placa', ''),
            getattr(b.conductor, 'nombres', ''),
            getattr(b.tipo_producto, 'nombre', ''),
            b.peso_ingreso_kg,
            b.peso_salida_kg,
            (b.peso_ingreso_kg or 0) - (b.peso_salida_kg or 0),
            b.guia_remision,
            b.guia_transporte,
            b.observaciones
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Guardar en memoria y asociar al reporte
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    reporte.archivo.save(f'reporte_balanza_{reporte.fecha_inicio}_{reporte.fecha_fin}.xlsx', ContentFile(output.read()), save=False)

@login_required
def balanza_pdf(request, balanza_id):
    balanza = get_object_or_404(Balanza, id=balanza_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ticket_balanza_{balanza.numero_guia_ticket}.pdf"'

    # Tamaño personalizado (21cm x 14cm)
    width, height = 21*cm, 14*cm
    c = canvas.Canvas(response, pagesize=(width, height))

    # Margen superior de 3.5cm
    y_start = height - 3.5*cm
    x_left = 2*cm
    x_right = 11.5*cm
    line_height = 18
    bold = 'Helvetica-Bold'
    normal = 'Helvetica'

    # --- Columna Izquierda (parte superior) ---
    y = y_start
    c.setFont(normal, 12)
    c.drawString(x_left, y, f"N' Ticket:  {balanza.numero_guia_ticket}")
    y -= line_height * 1.5

    # Proveedor
    proveedor_label = "Proveedor:"
    proveedor_nombre = balanza.facturador.razon_social
    max_prov_len = 38
    if len(proveedor_nombre) > max_prov_len:
        c.drawString(x_left, y, proveedor_label)
        c.setFont(bold, 12)
        c.drawString(x_left + 70, y, proveedor_nombre[:max_prov_len])
        y -= line_height
        c.drawString(x_left + 70, y, proveedor_nombre[max_prov_len:])
        c.setFont(normal, 12)
    else:
        c.drawString(x_left, y, proveedor_label)
        c.setFont(bold, 12)
        c.drawString(x_left + 70, y, proveedor_nombre)
        c.setFont(normal, 12)
    y -= line_height

    # Placa
    c.drawString(x_left, y, "Placa:")
    c.setFont(bold, 12)
    c.drawString(x_left + 45, y, balanza.vehiculo.placa)
    c.setFont(normal, 12)
    y -= line_height

    # Chofer
    chofer_label = "Chofer:"
    chofer_nombre = balanza.conductor.nombres
    max_chofer_len = 38
    if len(chofer_nombre) > max_chofer_len:
        c.drawString(x_left, y, chofer_label)
        c.drawString(x_left + 55, y, chofer_nombre[:max_chofer_len])
        y -= line_height
        c.drawString(x_left + 55, y, chofer_nombre[max_chofer_len:])
    else:
        c.drawString(x_left, y, f"{chofer_label} {chofer_nombre}")
    y -= line_height
    # Licencia (si existe)
    if hasattr(balanza.conductor, 'licencia') and balanza.conductor.licencia:
        c.drawString(x_left + 55, y, f"-licencia {balanza.conductor.licencia}")
        y -= line_height

    # Producto
    c.drawString(x_left, y, f"Producto:")
    c.drawString(x_left + 60, y, balanza.tipo_producto.nombre)
    y -= line_height

    # --- Columna Izquierda (parte inferior) ---
    # Preparamos los textos
    operador_label = "Operador:"
    operador_valor = "Planta"
    obs = "Obs:"
    obs_parts = []
    if balanza.lote_temporal:
        obs_parts.append(balanza.lote_temporal)
    if balanza.guia_remision:
        obs_parts.append(f"GRR:{balanza.guia_remision}")
    if balanza.guia_transporte:
        obs_parts.append(f"GRT:{balanza.guia_transporte}")
    # Agregar tipo de carga mineral en lugar de observaciones
    if balanza.tipo_empaque:
        if balanza.tipo_empaque == 'GRANEL':
            obs_parts.append("A GRANEL")
        elif balanza.tipo_empaque == 'SACOS' and balanza.cantidad_sacos:
            obs_parts.append(f"{balanza.cantidad_sacos} SACOS")
    obs_line = obs + ", ".join(obs_parts)
    max_obs_len = 38
    obs_lines = []
    while len(obs_line) > max_obs_len:
        obs_lines.append(obs_line[:max_obs_len])
        obs_line = obs_line[max_obs_len:]
    obs_lines.append(obs_line)
    empresa = "MINERA FIDAMI S.A"

    # Calculamos la altura total
    total_lines = 1 + len(obs_lines) + 1  # Operador + obs + empresa
    y_base = 2.5*cm + line_height * (total_lines - 1)
    y_op = y_base
    c.setFont(normal, 12)
    c.drawString(x_left, y_op, operador_label)
    c.setFont(bold, 12)
    c.drawString(x_left + 65, y_op, operador_valor)
    c.setFont(normal, 12)
    y_op -= line_height
    for l in obs_lines:
        c.drawString(x_left, y_op, l)
        y_op -= line_height
    c.drawString(x_left, y_op, empresa)

    # Línea horizontal al final de la hoja (ahora a 2.5cm)
    c.setLineWidth(1)
    c.line(x_left, 2.5*cm - 0.4*cm, width - x_left, 2.5*cm - 0.4*cm)

    # --- Columna Derecha ---
    y_r = y_start
    c.setFont(normal, 12)
    c.drawString(x_right, y_r, "INGRESO")
    y_r -= line_height
    c.setFont(normal, 11)
    c.drawString(x_right, y_r, f"{balanza.fecha_ingreso} {balanza.hora_ingreso}")
    y_r -= line_height
    c.setFont(bold, 15)
    c.drawString(x_right, y_r, "PESO   ")
    c.drawString(x_right + 60, y_r, f"{int(balanza.peso_ingreso_kg)} Kg")
    y_r -= line_height * 2

    c.setFont(normal, 12)
    c.drawString(x_right, y_r, "SALIDA")
    y_r -= line_height
    c.setFont(normal, 11)
    c.drawString(x_right, y_r, f"{balanza.fecha_salida} {balanza.hora_salida}")
    y_r -= line_height
    c.setFont(bold, 15)
    c.drawString(x_right, y_r, "PESO   ")
    c.drawString(x_right + 60, y_r, f"{int(balanza.peso_salida_kg)} Kg")
    y_r -= line_height * 2

    # Línea divisoria
    c.setLineWidth(1)
    c.line(x_right, y_r, width - 2*cm, y_r)
    y_r -= line_height

    # Peso cargado (neto)
    c.setFont(normal, 12)
    c.drawString(x_right, y_r, "PESO CARGADO (NETO)")
    y_r -= line_height * 1.5
    c.setFont(bold, 18)
    c.drawString(x_right, y_r, f"{int(balanza.peso_neto_kg)} Kg")

    c.save()
    return response

@login_required
def liquidacion_pdf(request, liquidacion_id):
    liquidacion = get_object_or_404(Liquidacion, pk=liquidacion_id)
    detalles = LiquidacionDetalle.objects.filter(liquidacion=liquidacion)

    # Cálculos
    subtotal = sum(det.lote.valorizacion.valorizacion_uds_tms or Decimal('0') for det in detalles)
    igv = subtotal * Decimal('0.18')
    total_valorizacion = subtotal + igv
    anticipo = sum(det.lote.valorizacion.anticipo or Decimal('0') for det in detalles)
    total_valorizacion_neto = total_valorizacion - anticipo
    detraccion = total_valorizacion_neto * Decimal('0.10')
    descuento_flete = sum(
        (det.lote.tmh or Decimal('0')) * (getattr(det.lote.valorizacion, 'pu_tmh_flete', Decimal('0')) or Decimal('0'))
        for det in detalles
    )
    monto_pagado = total_valorizacion_neto - detraccion - descuento_flete

    # PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="liquidacion_{liquidacion.id}.pdf"'
    w, h = landscape(A4)
    c = canvas.Canvas(response, pagesize=(w, h))

    # Marca de agua (logo)
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'logo.png')
    if os.path.exists(logo_path):
        logo = ImageReader(logo_path)
        c.saveState()
        c.translate(w/2-7*cm, h/2-7*cm)
        c.setFillAlpha(0.08)
        c.drawImage(logo, 0, 0, width=14*cm, height=14*cm, mask='auto')
        c.restoreState()

    # Título
    c.setFont('Helvetica-Bold', 20)
    c.drawCentredString(w/2, h-2*cm, "LIQUIDACION DE MINERAL")

    # Datos generales (alineados y separados)
    c.setFont('Helvetica', 9)
    y = h-3*cm
    x_label = 2*cm
    x_value = x_label + 3.5*cm  # separación más amplia
    datos_generales = [
        ("Fecha de Emisión:", liquidacion.fecha_creacion.strftime('%d de %B de %Y')),
        ("Proveedor:", liquidacion.proveedor.razon_social),
        ("RUC:", liquidacion.proveedor.ruc),
        ("C. INGEMMET:", detalles.first().lote.codigo_ingemmet.codigo_ingemmet if detalles.first() else ''),
        ("Procedencia:", detalles.first().lote.procedencia if detalles.first() else ''),
        ("Tipo Mineral:", getattr(detalles.first().lote.tipo_producto, 'nombre', '') if detalles else ''),
        ("Guía Remitente:", getattr(detalles.first().lote, 'guia_remision', '') if detalles else ''),
        ("Guía Transporte:", getattr(detalles.first().lote, 'guia_transporte', '') if detalles else ''),
    ]
    for label, value in datos_generales:
        c.setFont('Helvetica-Bold', 9)
        c.drawString(x_label, y, label)
        c.setFont('Helvetica', 9)
        c.drawString(x_value, y, str(value))
        y -= 0.5*cm

    # Tabla de lotes
    y_tabla = h-7.5*cm  # espacio ligeramente reducido después de los datos generales
    # Ancho útil de la hoja (margen de 2cm a cada lado)
    ancho_util = w - 4*cm
    # Proporciones ajustadas para columnas (más angostas las cortas, más espacio a las largas)
    proporciones = [0.08, 0.11, 0.07, 0.07, 0.08, 0.09, 0.07, 0.10, 0.08, 0.05, 0.05, 0.15]
    col_widths = [ancho_util * p for p in proporciones]
    headers = [
        "N° LOTE", "Fecha de Recepción", "TMH", "% Hum.", "TMS", "LEY Oz/Tc AU", "Recup. %", "P.I.O US$ Oz", "Maquila X TMS", "PENA", "Factor", "Valorización Total / US$"
    ]
    c.setFont('Helvetica-Bold', 8)
    x = 2*cm
    for i, htxt in enumerate(headers):
        c.setFillColorRGB(0.36, 0.51, 0.13)  # verde oscuro
        c.rect(x-0.1*cm, y_tabla-0.15*cm, col_widths[i], 0.7*cm, fill=1, stroke=0)
        c.setFillColor(colors.white)
        # Dividir el título en dos líneas si es largo
        if ' ' in htxt:
            partes = htxt.split(' ', 1)
            c.drawCentredString(x + col_widths[i]/2, y_tabla+0.25*cm, partes[0])
            c.drawCentredString(x + col_widths[i]/2, y_tabla-0.05*cm, partes[1])
        else:
            c.drawCentredString(x + col_widths[i]/2, y_tabla+0.1*cm, htxt)
        x += col_widths[i]
    c.setFillColor(colors.black)
    c.setFont('Helvetica', 8)
    y_tabla -= 0.7*cm
    idx_numericas = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
    for det in detalles:
        x = 2*cm
        fila = [
            det.lote.codigo_lote,
            det.lote.fecha_ingreso.strftime('%d/%m/%y'),
            f"{det.lote.tmh:,.3f}",
            f"{det.lote.ley.porcentaje_h2o:,.2f}",
            f"{det.lote.ley.tms:,.3f}",
            f"{det.lote.ley.ley_onz_tc:,.3f}",
            f"{det.lote.ley.porcentaje_recuperacion:,.0f}",
            f"{det.lote.costo.pio_us_onza:,.2f}",
            f"{det.lote.costo.maquila:,.2f}",
            "10.00",
            "1.10231",
            f"{det.lote.valorizacion.valorizacion_uds_tms:,.2f}",
        ]
        for i, val in enumerate(fila):
            c.setFillColor(colors.black)
            if i == len(fila) - 1:
                c.drawRightString(x + col_widths[i] - 2, y_tabla, str(val))
            else:
                c.drawCentredString(x + col_widths[i]/2, y_tabla, str(val))
            x += col_widths[i]
        y_tabla -= 0.6*cm

    # Subtotales y totales (alineados, separados y con fuentes más pequeñas)
    y_tot = y_tabla - 0.5*cm
    x_titulo = w-15*cm
    x_valor = w-2*cm
    ancho_fila = x_valor - x_titulo
    c.setFont('Helvetica-Bold', 9)
    c.drawString(x_titulo, y_tot, "SUB TOTAL:")
    c.setFont('Helvetica-Bold', 9)
    c.drawRightString(x_valor, y_tot, f"{subtotal:,.2f}")
    y_tot -= 0.5*cm
    c.setFont('Helvetica-Bold', 9)
    c.drawString(x_titulo, y_tot, "IGV 18%:")
    c.drawRightString(x_valor, y_tot, f"{igv:,.2f}")
    y_tot -= 0.5*cm
    c.setLineWidth(0.5)
    c.setStrokeColor(colors.lightgrey)
    c.line(x_titulo, y_tot+0.3*cm, x_valor, y_tot+0.3*cm)
    c.setFont('Helvetica-Bold', 10.5)
    c.setFillColor(colors.black)
    c.drawString(x_titulo, y_tot, "TOTAL VALORIZACIÓN DE MINERAL:")
    c.drawRightString(x_valor, y_tot, f"{total_valorizacion:,.2f}")
    y_tot -= 0.7*cm
    c.setLineWidth(0.5)
    c.line(x_titulo, y_tot+0.5*cm, x_valor, y_tot+0.5*cm)
    c.setFont('Helvetica-Bold', 9)
    c.drawString(x_titulo, y_tot, "(-) ANTICIPO:")
    c.drawRightString(x_valor, y_tot, f"{anticipo:,.2f}")
    y_tot -= 0.5*cm
    c.setFont('Helvetica-Bold', 10.5)
    c.drawString(x_titulo, y_tot, "TOTAL VALORIZACIÓN:")
    c.drawRightString(x_valor, y_tot, f"{total_valorizacion_neto:,.2f}")
    y_tot -= 0.7*cm
    c.setLineWidth(0.5)
    c.line(x_titulo, y_tot+0.5*cm, x_valor, y_tot+0.5*cm)
    c.setFont('Helvetica-Bold', 9)
    c.drawString(x_titulo, y_tot, "(-) DETRACCIÓN (10%):")
    c.drawRightString(x_valor, y_tot, f"{detraccion:,.2f}")
    y_tot -= 0.5*cm
    c.drawString(x_titulo, y_tot, "(-) DESCUENTO FLETE:")
    c.drawRightString(x_valor, y_tot, f"{descuento_flete:,.2f}")
    y_tot -= 0.7*cm
    # Fondo azul para el monto a pagar
    alto_fila = 0.7*cm
    c.setFillColorRGB(0.8, 0.89, 1)  # azul claro
    c.rect(x_titulo, y_tot-alto_fila*0.2, ancho_fila, alto_fila, fill=1, stroke=0)
    c.setFillColor(colors.black)
    c.setFont('Helvetica-Bold', 11)
    c.drawString(x_titulo+0.2*cm, y_tot+0.2*cm, "MONTO A PAGAR:")
    c.setFont('Helvetica-Bold', 12)
    c.drawRightString(x_valor-0.2*cm, y_tot+0.2*cm, f"US$ {monto_pagado:,.2f}")
    c.setFillColor(colors.black)
    y_tot -= 0.9*cm

    # Observaciones y datos bancarios con separación de etiqueta y valor
    y_obs = y_tabla - 1.2*cm
    x_label = 2*cm
    x_value = x_label + 4*cm
    primer_lote = detalles.first().lote if detalles else None
    banco = primer_lote.valorizacion.banco if primer_lote and hasattr(primer_lote, 'valorizacion') else ''
    cuenta = primer_lote.valorizacion.cuenta if primer_lote and hasattr(primer_lote, 'valorizacion') else ''
    c.setFont('Helvetica-Bold', 10)
    c.drawString(x_label, y_obs, "OBSERVACIÓN:")
    c.setFont('Helvetica', 10)
    c.drawString(x_label, y_obs-0.5*cm, "TRANSPORTISTA:")
    c.drawString(x_value, y_obs-0.5*cm, f"{getattr(detalles.first().lote.transportista, 'razon_social', '') if detalles else ''}")
    c.drawString(x_label, y_obs-1*cm, "PLACA:")
    c.drawString(x_value, y_obs-1*cm, f"{getattr(detalles.first().lote.vehiculo, 'placa', '') if detalles else ''}")
    c.drawString(x_label, y_obs-1.5*cm, "BANCO:")
    c.drawString(x_value, y_obs-1.5*cm, banco)
    c.drawString(x_label, y_obs-2*cm, "CUENTA:")
    c.drawString(x_value, y_obs-2*cm, cuenta)

    # Línea para firma y nombre del proveedor a la izquierda, más cerca del borde inferior
    y_firma = 1.5*cm
    c.setLineWidth(1)
    c.line(2*cm, y_firma+0.8*cm, 12*cm, y_firma+0.8*cm)
    c.setFont('Helvetica', 10)
    c.drawString(2*cm, y_firma+0.3*cm, liquidacion.proveedor.razon_social)

    c.save()
    return response

# Función para generar el reporte de campañas en Excel
def generar_reporte_campanas_excel(reporte):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Campañas'

    # Encabezados
    headers = [
        'Nombre', 'Descripción', 'Fecha Inicio', 'Fecha Fin', 'Estado', 'Lotes Asociados', 'Total TMH', 'Total TMS Real'
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Filtrar campañas por fechas
    campanas = Campana.objects.filter(
        fecha_inicio__gte=reporte.fecha_inicio,
        fecha_fin__lte=reporte.fecha_fin
    ).order_by('fecha_inicio')

    for campana in campanas:
        lotes = Lote.objects.filter(campanalote__campana=campana).select_related('costo')
        lotes_str = ', '.join([lote.codigo_lote for lote in lotes])
        total_tmh = sum(lote.tmh or 0 for lote in lotes)
        total_tms = sum(getattr(lote.costo, 'tms_real', 0) or 0 for lote in lotes)
        ws.append([
            campana.nombre,
            campana.descripcion or '',
            campana.fecha_inicio.strftime('%Y-%m-%d'),
            campana.fecha_fin.strftime('%Y-%m-%d'),
            campana.get_estado_display(),
            lotes_str,
            total_tmh,
            total_tms
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Guardar en memoria y asociar al reporte
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    reporte.archivo.save(f'reporte_campanas_{reporte.fecha_inicio}_{reporte.fecha_fin}.xlsx', ContentFile(output.read()), save=False)
